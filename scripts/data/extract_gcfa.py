#!/usr/bin/env python3.11
"""
Extract GCFA local-church statistical tables (2000-2024) from the service-ticket
workbook into clean, normalized artifacts keyed by the stable GCFA church number.

Inputs : the GCFA service-ticket workbook (one sheet per year: Headers + 2000..2024).
         Default path is looked up by CONFERENCE_SLUG in DEFAULT_WORKBOOKS below
         (Rio Texas: Service Ticket #726954 - Wilson Pruitt.xlsx); pass an explicit
         path for any conference without a registered default.
Outputs (scripts/data/gcfa/):
  codebook.json      canonical stat fields  {code, label, question, table, category}
  churches.json      one record per GCFA number (latest-known identity + history span)
  church_stats.csv   long format: gcfa_number,data_year,conference,field_code,value_numeric,value_text
  qa_report.txt      coverage / sanity summary

Run: /usr/local/bin/python3.11 scripts/data/extract_gcfa.py [path-to-xlsx]
"""
import openpyxl, re, json, csv, sys, os, datetime
from collections import defaultdict, Counter

# GCFA's local-church statistical-table format is nationwide, so this extractor is
# ~already the universal Tier-A adapter (see
# ~/wroot-labs/notes/conference-atlas-architecture.md §4) — only the default
# workbook path is per-conference. Add an entry here per new conference, mirroring
# the same CONFERENCE_SLUG-keyed-registry pattern as src/lib/conference.ts.
DEFAULT_WORKBOOKS = {
    "rio-texas": os.path.expanduser("~/Downloads/Service Ticket #726954 - Wilson Pruitt.xlsx"),
}
CONFERENCE_SLUG = os.environ.get("CONFERENCE_SLUG", "rio-texas")

if len(sys.argv) > 1:
    XLSX = sys.argv[1]
elif CONFERENCE_SLUG in DEFAULT_WORKBOOKS:
    XLSX = DEFAULT_WORKBOOKS[CONFERENCE_SLUG]
else:
    raise SystemExit(
        f"No default workbook registered for conference '{CONFERENCE_SLUG}' — "
        f"pass the path explicitly: extract_gcfa.py <path-to-xlsx>")
OUT = os.path.join(os.path.dirname(__file__), "gcfa")
os.makedirs(OUT, exist_ok=True)


def norm(h):
    """Normalize a header to a canonical comparison key."""
    if h is None:
        return None
    s = str(h).strip()
    s = re.sub(r'_(19|20)\d\d$', '', s)   # strip _YYYY suffix
    s = re.sub(r'_\d\d$', '', s)          # strip _YY suffix
    s = re.sub(r'[^A-Za-z0-9]', '', s).upper()
    return s or None


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    return v


CONF_CANON = {
    'RIO TEXAS': 'Rio Texas', 'RIOTEXAS': 'Rio Texas',
    'RIO GRANDE': 'Rio Grande', 'RIOGRANDE': 'Rio Grande',
    'SOUTHWEST TEXAS': 'Southwest Texas', 'SOUTHWESTTEXAS': 'Southwest Texas',
}


def canon_conf(v):
    if v is None:
        return None
    return CONF_CANON.get(str(v).strip().upper(), str(v).strip())


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('$', '')
    if s in ('', '-', 'N/A', 'NA', '.'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---- identity column aliases (normalized) ----
IDENTITY = {
    'gcfa_number':       ['GCNO', 'CHURCHNO', 'CHURCH'],
    'gcfa_id':           ['GCFAID'],
    'church_name':       ['CHURCHNAME'],
    'conference_no':     ['CONFERENCENO', 'CONFNO', 'CONF'],
    'conference_name':   ['CONFERENCENAME', 'CONFNAME'],
    'district_no':       ['DISTRISTNO', 'DISTNO', 'DISTRICTNO'],
    'district_name':     ['DISTNAME', 'DISTRICTNAME', 'DISTRICT'],
    'city':              ['CITY', 'LOCATION'],
    'state':             ['STATE', 'STATEPROVINCE'],
    'zip':               ['ZIPCODE', 'ZIP'],
    'county_no':         ['COUNTYNO'],
    'county_name':       ['COUNTYNAME'],
    'church_status':     ['CHURCHSTATUS'],
    'congregation_type': ['CONGREGATIONTYPE', 'CHURCHTYPE'],
    'church_ethnicity':  ['CHURCHETHNICITY'],
    'ein':               ['CHURCHEIN', 'EIN'],
    'charge_no':         ['CHARGENO'],
    'charge_name':       ['CHARGENAME'],
    'address1':          ['ADDRESS1'],
    'closed_date':       ['CLOSEDDATE'],
    'closure_reason':    ['CLOSUREREASON'],
}
# normalized identity keys we should NOT treat as stat fields
IDENTITY_KEYS = set()
for alts in IDENTITY.values():
    IDENTITY_KEYS.update(alts)
# extra non-stat columns to ignore in old layout
IGNORE = {'SORTNAME', 'DISPLAYNAME', 'PARISHNO', 'PARISH', 'LOCATIONNO',
          'CHARGESORTNAME', 'CHURSORTNM', 'ISCHARGE', 'CHURCHCOUNT', 'LABELID',
          'PARTNO', 'TITLE', 'LASTNAME', 'FIRSTNAME', 'MIDDLENAME', 'SUFFIX',
          'STATUS', 'CLERGYNAME', 'CLERGYETHNICITY', 'PASTORNAME',
          'PASTORETHNICITY', 'PASTORSTATUS', 'DISTRICTSORTNAME',
          'REOPENEDDATE', 'CHURCHESINLITIGATION', 'ADDRESS2'}


def build_idmap(hdr):
    """Map field-name -> column index for one sheet's header row."""
    nh = [norm(h) for h in hdr]
    idx = {}
    for field, alts in IDENTITY.items():
        for a in alts:
            if a in nh:
                idx[field] = nh.index(a)
                break
    return idx, nh


# ---------------------------------------------------------------- codebook
def load_codebook(wb):
    ws = wb['Headers']
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    ci = {str(h).strip(): i for i, h in enumerate(hdr) if h}
    code_i = ci.get('Column Heading')
    q_i = ci.get('Question')
    tbl_i = ci.get('Table')
    cb = {}
    for r in rows[1:]:
        code = clean(r[code_i]) if code_i is not None else None
        if not code:
            continue
        ncode = norm(code)
        cb[ncode] = {
            'code': ncode,
            'raw_code': code,
            'question': clean(r[q_i]) if q_i is not None else None,
            'table': clean(r[tbl_i]) if tbl_i is not None else None,
        }
    return cb


def classify_unit(code, question):
    q = (question or '').lower()
    c = (code or '').upper()
    money = ['paid', 'salary', 'compensation', 'housing', 'pension', 'expend',
             'expense', 'debt', 'capital', 'apportion', 'benevolence', 'fund',
             'contribut', 'income', 'cash', 'receipt', 'offering', 'amount', '$',
             'dollars', 'budget', 'asset']
    money_codes = ('SAL', 'CASH', 'TOT', 'PAID', 'EXP', 'FUND', 'APPOR', 'BEN',
                   'GCFA', 'CONF', 'DIST', 'DEBT', 'OFF', 'PASTCASH', 'ASSCCASH')
    if any(w in q for w in money):
        return 'usd'
    if c.endswith('PD') or c.endswith('PAID') or any(m in c for m in ('SAL', 'CASH', 'EXPEND', 'APPOR', 'BENEV')):
        return 'usd'
    return 'count'


def categorize(question, table):
    q = (question or '').lower()
    if any(w in q for w in ['member', 'baptiz', 'profess', 'constituen', 'attendance', 'worship', 'formation', 'small group', 'vbs', 'confirm']):
        return 'membership'
    if any(w in q for w in ['asian', 'african', 'hispanic', 'native', 'pacific', 'white', 'multi racial', 'multiracial', 'female', 'male', 'non-binary', 'ethnicity', 'gender']):
        return 'demographics'
    if any(w in q for w in ['apportion', 'benevolence', 'mission', 'general church']):
        return 'apportionments'
    if any(w in q for w in ['salary', 'compensation', 'housing', 'pension', 'paid', 'expense', 'expend', 'debt', 'capital', 'building', 'operating']):
        return 'finance'
    if any(w in q for w in ['received', 'income', 'contribut', 'fund', 'tithe', 'offering', 'receipt']):
        return 'giving'
    return 'other'


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    codebook = load_codebook(wb)
    for c in codebook.values():
        c['category'] = categorize(c['question'], c['table'])

    years = sorted(int(s) for s in wb.sheetnames if s != 'Headers')
    churches = {}                 # gcfa -> identity record
    stats_rows = []               # long-format
    used_codes = {}               # code -> max data_year seen (for last_seen)
    used_codes_first = {}         # code -> min data_year
    unknown_codes = Counter()
    qa = []

    for yr in years:
        ws = wb[str(yr)]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        idx, nh = build_idmap(hdr)
        if 'gcfa_number' not in idx:
            qa.append(f"{yr}: NO GCFA KEY COLUMN — skipped")
            continue
        # which columns are stat fields?
        stat_cols = []
        for ci, code in enumerate(nh):
            if code is None:
                continue
            if code in IDENTITY_KEYS or code in IGNORE:
                continue
            stat_cols.append((ci, code))
        nrows = 0
        nstats = 0
        confs = Counter()
        for r in it:
            gid = clean(r[idx['gcfa_number']])
            if gid is None:
                continue
            gid = str(int(gid)) if isinstance(gid, (int, float)) else str(gid).strip()
            nrows += 1
            conf = canon_conf(clean(r[idx['conference_name']])) if 'conference_name' in idx else None
            confs[conf] += 1
            # identity (keep latest year seen = overwrite, years ascending)
            rec = churches.setdefault(gid, {'gcfa_number': gid, 'first_year': yr, 'last_year': yr,
                                            'conferences': set(), 'years': []})
            rec['last_year'] = yr
            rec['years'].append(yr)
            if conf:
                rec['conferences'].add(conf)
            for field in IDENTITY:
                if field in idx and field != 'gcfa_number':
                    val = clean(r[idx[field]])
                    if val is not None:
                        rec[field] = val
            # stats
            for ci, code in stat_cols:
                raw = r[ci] if ci < len(r) else None
                if raw is None:
                    continue
                canon = code if code in codebook else None
                if canon is None:
                    unknown_codes[code] += 1
                    # still record under raw normalized code
                used_codes[code] = max(used_codes.get(code, yr), yr)
                used_codes_first[code] = min(used_codes_first.get(code, yr), yr)
                num = to_num(raw)
                stats_rows.append({
                    'gcfa_number': gid, 'data_year': yr, 'conference': conf,
                    'field_code': code,
                    'value_numeric': num,
                    'value_text': None if num is not None else str(raw).strip(),
                })
                nstats += 1
        qa.append(f"{yr}: {nrows} churches, {nstats} stat cells, {len(stat_cols)} stat cols, conf={dict(confs)}")

    # finalize church records
    out_churches = []
    for gid, rec in sorted(churches.items()):
        rec['conferences'] = sorted(rec['conferences'])
        rec['years'] = sorted(set(rec['years']))
        rec['n_years'] = len(rec['years'])
        out_churches.append(rec)

    # build complete field list (every code that appears in the data, codebook-enriched)
    fields = []
    for code in sorted(used_codes):
        cb = codebook.get(code)
        q = cb['question'] if cb else None
        fields.append({
            'code': code,
            'label': q or code,
            'question': q,
            'category': cb['category'] if cb else 'other',
            'unit': classify_unit(code, q),
            'table_no': str(cb['table']) if cb and cb['table'] is not None else None,
            'in_codebook': cb is not None,
            'first_seen_year': used_codes_first[code],
            'last_seen_year': used_codes[code],
        })

    # write outputs
    with open(os.path.join(OUT, 'fields.json'), 'w') as f:
        json.dump(fields, f, indent=2, ensure_ascii=False)
    with open(os.path.join(OUT, 'codebook.json'), 'w') as f:
        json.dump(list(codebook.values()), f, indent=2, ensure_ascii=False)
    with open(os.path.join(OUT, 'churches.json'), 'w') as f:
        json.dump(out_churches, f, indent=2, ensure_ascii=False)
    with open(os.path.join(OUT, 'church_stats.csv'), 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['gcfa_number', 'data_year', 'conference',
                                          'field_code', 'value_numeric', 'value_text'])
        w.writeheader()
        w.writerows(stats_rows)
    # JSONL for the Node importer (no CSV quoting ambiguity)
    with open(os.path.join(OUT, 'church_stats.jsonl'), 'w') as f:
        for r in stats_rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    # QA report
    lines = []
    lines.append(f"GCFA EXTRACT QA — {XLSX}")
    lines.append(f"years: {years[0]}..{years[-1]}  ({len(years)} sheets)")
    lines.append(f"codebook fields: {len(codebook)}")
    lines.append(f"unique churches (by GCFA #): {len(out_churches)}")
    lines.append(f"total stat cells: {len(stats_rows)}")
    lines.append("")
    lines.append("PER-YEAR COVERAGE:")
    lines += ["  " + l for l in qa]
    lines.append("")
    lines.append(f"UNKNOWN STAT CODES (not in codebook): {len(unknown_codes)}")
    for code, n in unknown_codes.most_common(40):
        lines.append(f"  {code}: {n}")
    # distribution of how many years churches appear
    span = Counter(c['n_years'] for c in out_churches)
    lines.append("")
    lines.append("CHURCH LONGEVITY (n_years appearing): " + str(dict(sorted(span.items()))))
    rep = "\n".join(lines)
    with open(os.path.join(OUT, 'qa_report.txt'), 'w') as f:
        f.write(rep)
    print(rep)


if __name__ == '__main__':
    main()
